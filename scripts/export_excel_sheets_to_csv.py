#!/usr/bin/env python3
"""
통합 엑셀의 시트를 각 Terraform stack의 data/*.csv로 내보내는 스크립트입니다.

필요 패키지:
  pip install openpyxl

사용:
  python3 scripts/export_excel_sheets_to_csv.py --xlsx terraform_root_30users_design.xlsx --root .
"""
import argparse
import csv
from pathlib import Path
from openpyxl import load_workbook

SHEET_TO_CSV = {
    "00_users": "00-entra-org/data/users.csv",
    "00_groups": "00-entra-org/data/groups.csv",
    "00_group_memberships": "00-entra-org/data/group_memberships.csv",
    "00_azure_rbac": "00-entra-org/data/azure_rbac.csv",
    "00_aks_rbac": "00-entra-org/data/aks_rbac.csv",
    "10_resource_groups": "10-network/data/resource_groups.csv",
    "10_vnets": "10-network/data/vnets.csv",
    "10_subnets": "10-network/data/subnets.csv",
    "10_nsgs": "10-network/data/nsgs.csv",
    "10_nsg_rules": "10-network/data/nsg_rules.csv",
    "20_resource_groups": "20-aks/data/resource_groups.csv",
    "20_acr": "20-aks/data/acr.csv",
    "20_monitor": "20-aks/data/monitor.csv",
    "20_aks_clusters": "20-aks/data/aks_clusters.csv",
    "20_nodepools": "20-aks/data/nodepools.csv",
    "20_aks_azure_rbac": "20-aks/data/aks_azure_rbac.csv",
    "30_namespaces": "30-aks-rbac/data/namespaces.csv",
    "30_roles": "30-aks-rbac/data/roles.csv",
    "30_rolebindings": "30-aks-rbac/data/rolebindings.csv",
}

def sheet_rows(ws):
    values = []
    for row in ws.iter_rows(values_only=True):
        row_values = ["" if v is None else v for v in row]
        if any(str(v).strip() for v in row_values):
            values.append(row_values)
    if not values:
        return []
    # trim trailing empty columns based on header
    header = values[0]
    last = 0
    for i, v in enumerate(header):
        if str(v).strip():
            last = i
    return [r[:last+1] for r in values]

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True)
    parser.add_argument("--root", default=".")
    args = parser.parse_args()

    root = Path(args.root).resolve()
    wb = load_workbook(args.xlsx, data_only=False)

    for sheet_name, rel_path in SHEET_TO_CSV.items():
        if sheet_name not in wb.sheetnames:
            print(f"SKIP: sheet not found: {sheet_name}")
            continue
        rows = sheet_rows(wb[sheet_name])
        out = root / rel_path
        out.parent.mkdir(parents=True, exist_ok=True)
        with out.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerows(rows)
        print(f"OK: {sheet_name} -> {out}")

if __name__ == "__main__":
    main()
